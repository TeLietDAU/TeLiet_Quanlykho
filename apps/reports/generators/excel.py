from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


class ExcelTableMixin:
    def build_excel_table(self, title, headers, rows, meta=None):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Report'

        row_cursor = 1
        sheet.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=max(len(headers), 1))
        sheet.cell(row=row_cursor, column=1, value=title)
        sheet.cell(row=row_cursor, column=1).font = Font(size=14, bold=True)
        sheet.cell(row=row_cursor, column=1).alignment = Alignment(horizontal='center')

        row_cursor += 1
        for line in (meta or []):
            sheet.cell(row=row_cursor, column=1, value=line)
            row_cursor += 1

        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row_cursor, column=column, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(fill_type='solid', fgColor='1F4E78')
            cell.alignment = Alignment(horizontal='center')

        row_cursor += 1
        for row_values in rows:
            for column, value in enumerate(row_values, start=1):
                sheet.cell(row=row_cursor, column=column, value=value)
            row_cursor += 1

        for index, _ in enumerate(headers, start=1):
            sheet.column_dimensions[chr(64 + index)].width = 18

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
