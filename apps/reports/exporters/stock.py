from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from django.http import HttpResponse
from django.utils import timezone
import os

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.reports.utils import format_report_number

class StockReportExporter:
    @staticmethod
    def export_excel(rows, totals, from_date, to_date, category_label, generated_by):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Bao cao ton kho'

        title = 'BAO CAO TON KHO THEO THOI GIAN'

        sheet.merge_cells('A1:H1')
        sheet['A1'] = title
        sheet['A1'].font = Font(size=14, bold=True)
        sheet['A1'].alignment = Alignment(horizontal='center')

        sheet['A2'] = f'Tu ngay: {from_date.strftime("%d/%m/%Y")}'
        sheet['A3'] = f'Den ngay: {to_date.strftime("%d/%m/%Y")}'
        sheet['A4'] = f'Danh muc: {category_label}'
        sheet['A5'] = f'Xuat luc: {timezone.localtime().strftime("%d/%m/%Y %H:%M:%S")}'
        sheet['A6'] = f'Nguoi xuat: {generated_by}'

        headers = [
            'STT',
            'San pham',
            'Danh muc',
            'Don vi',
            'Ton dau ky',
            'Nhap trong ky',
            'Xuat trong ky',
            'Ton cuoi ky',
        ]

        header_row = 7
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(fill_type='solid', fgColor='1F4E78')
            cell.alignment = Alignment(horizontal='center')

        number_format = '#,##0.##'
        start_data_row = header_row + 1
        current_row = start_data_row
        for index, row in enumerate(rows, start=1):
            sheet.cell(row=current_row, column=1, value=index)
            sheet.cell(row=current_row, column=2, value=row['product'].name)
            sheet.cell(row=current_row, column=3, value=row['product'].category.name if row['product'].category else '')
            sheet.cell(row=current_row, column=4, value=row['product'].base_unit)

            opening_cell = sheet.cell(row=current_row, column=5, value=float(row['opening']))
            import_cell = sheet.cell(row=current_row, column=6, value=float(row['import_qty']))
            export_cell = sheet.cell(row=current_row, column=7, value=float(row['export_qty']))
            closing_cell = sheet.cell(row=current_row, column=8, value=float(row['closing']))

            opening_cell.number_format = number_format
            import_cell.number_format = number_format
            export_cell.number_format = number_format
            closing_cell.number_format = number_format

            current_row += 1

        total_row = current_row
        sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
        total_title_cell = sheet.cell(row=total_row, column=1, value='Tong cong')
        total_title_cell.font = Font(bold=True)
        total_title_cell.alignment = Alignment(horizontal='right')

        total_opening = sheet.cell(row=total_row, column=5, value=float(totals['opening']))
        total_import = sheet.cell(row=total_row, column=6, value=float(totals['import_qty']))
        total_export = sheet.cell(row=total_row, column=7, value=float(totals['export_qty']))
        total_closing = sheet.cell(row=total_row, column=8, value=float(totals['closing']))

        for total_cell in (total_opening, total_import, total_export, total_closing):
            total_cell.number_format = number_format
            total_cell.font = Font(bold=True)

        column_widths = {
            'A': 7, 'B': 32, 'C': 20, 'D': 12, 'E': 15, 'F': 15, 'G': 15, 'H': 15,
        }
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width

        filename = f'bao_cao_ton_kho_{from_date.isoformat()}_{to_date.isoformat()}.xlsx'
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        workbook.save(response)
        return response

    @staticmethod
    def export_pdf(rows, totals, from_date, to_date, category_label, generated_by):
        filename = f'bao_cao_ton_kho_{from_date.isoformat()}_{to_date.isoformat()}.pdf'

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(A4),
            leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20,
        )

        styles = getSampleStyleSheet()
        regular_font = 'Helvetica'
        bold_font = 'Helvetica-Bold'

        regular_font_path = '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'
        bold_font_path = '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'
        if os.path.exists(regular_font_path) and os.path.exists(bold_font_path):
            pdfmetrics.registerFont(TTFont('DejaVuSans', regular_font_path))
            pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', bold_font_path))
            regular_font = 'DejaVuSans'
            bold_font = 'DejaVuSans-Bold'

        title_style = ParagraphStyle(
            'ReportTitle', parent=styles['Heading1'],
            alignment=TA_CENTER, fontName=bold_font, fontSize=16, spaceAfter=10,
        )
        body_style = ParagraphStyle(
            'ReportBody', parent=styles['Normal'],
            fontName=regular_font, fontSize=10,
        )

        story = [
            Paragraph('BAO CAO TON KHO THEO THOI GIAN', title_style),
            Paragraph(f'Tu ngay: {from_date.strftime("%d/%m/%Y")}', body_style),
            Paragraph(f'Den ngay: {to_date.strftime("%d/%m/%Y")}', body_style),
            Paragraph(f'Danh muc: {category_label}', body_style),
            Paragraph(f'Xuat luc: {timezone.localtime().strftime("%d/%m/%Y %H:%M:%S")}', body_style),
            Paragraph(f'Nguoi xuat: {generated_by}', body_style),
            Spacer(1, 10),
        ]

        def fmt(value):
            return format_report_number(value)

        table_data = [[
            'STT', 'San pham', 'Danh muc', 'Don vi', 'Ton dau ky', 'Nhap trong ky', 'Xuat trong ky', 'Ton cuoi ky',
        ]]

        for index, row in enumerate(rows, start=1):
            table_data.append([
                str(index),
                str(row['product'].name),
                str(row['product'].category.name if row['product'].category else ''),
                str(row['product'].base_unit),
                fmt(row['opening']),
                fmt(row['import_qty']),
                fmt(row['export_qty']),
                fmt(row['closing']),
            ])

        table_data.append([
            '', 'TONG CONG', '', '',
            fmt(totals['opening']), fmt(totals['import_qty']), fmt(totals['export_qty']), fmt(totals['closing']),
        ])

        table = Table(
            table_data,
            repeatRows=1,
            colWidths=[30, 180, 120, 70, 85, 85, 85, 85],
        )
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), bold_font),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
            ('FONTNAME', (0, 1), (-1, -2), regular_font),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F3F6FA')),
            ('FONTNAME', (0, -1), (-1, -1), bold_font),
        ])
        table.setStyle(table_style)

        story.append(table)
        doc.build(story)
        return response
