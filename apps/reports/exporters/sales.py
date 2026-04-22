from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from django.http import HttpResponse
from django.utils import timezone
from decimal import Decimal
import os

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.reports.utils import format_report_number, get_user_display_name
from apps.order.models import SalesOrder

class SalesOrderExporter:
    @staticmethod
    def export_excel(orders, status_filter, search_query, from_date, to_date, generated_by):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Bao cao don hang'

        status_labels = dict(SalesOrder.STATUS_CHOICES)
        status_label = status_labels.get(status_filter, status_filter) if status_filter else 'Tat ca trang thai'
        search_label = search_query if search_query else 'Khong co'

        sheet.merge_cells('A1:J1')
        sheet['A1'] = 'BAO CAO DON HANG'
        sheet['A1'].font = Font(size=14, bold=True)
        sheet['A1'].alignment = Alignment(horizontal='center')

        sheet['A2'] = f'Tu ngay: {from_date.strftime("%d/%m/%Y")}'
        sheet['A3'] = f'Den ngay: {to_date.strftime("%d/%m/%Y")}'
        sheet['A4'] = f'Trang thai: {status_label}'
        sheet['A5'] = f'Tu khoa tim kiem: {search_label}'
        sheet['A6'] = f'Xuat luc: {timezone.localtime().strftime("%d/%m/%Y %H:%M:%S")}'
        sheet['A7'] = f'Nguoi xuat: {generated_by}'

        headers = [
            'STT', 'Ma don', 'Khach hang', 'So dien thoai', 'Nhan vien tao',
            'Trang thai', 'Ngay tao', 'So dong SP', 'Tong so luong', 'Tong tien',
        ]

        header_row = 9
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(fill_type='solid', fgColor='1F4E78')
            cell.alignment = Alignment(horizontal='center')

        number_format = '#,##0.##'
        total_quantity = Decimal('0')
        total_amount = Decimal('0')
        current_row = header_row + 1

        for index, order in enumerate(orders, start=1):
            order_items = list(order.items.all())
            item_count = len(order_items)
            quantity_sum = sum((item.quantity for item in order_items), Decimal('0'))
            amount_sum = sum((item.subtotal for item in order_items), Decimal('0'))

            total_quantity += quantity_sum
            total_amount += amount_sum

            sheet.cell(row=current_row, column=1, value=index)
            sheet.cell(row=current_row, column=2, value=order.order_code)
            sheet.cell(row=current_row, column=3, value=order.customer_name)
            sheet.cell(row=current_row, column=4, value=order.customer_phone or '')
            sheet.cell(row=current_row, column=5, value=get_user_display_name(order.created_by))
            sheet.cell(row=current_row, column=6, value=order.get_status_display())
            sheet.cell(row=current_row, column=7, value=timezone.localtime(order.created_at).strftime('%d/%m/%Y %H:%M'))
            sheet.cell(row=current_row, column=8, value=item_count)

            quantity_cell = sheet.cell(row=current_row, column=9, value=float(quantity_sum))
            amount_cell = sheet.cell(row=current_row, column=10, value=float(amount_sum))
            quantity_cell.number_format = number_format
            amount_cell.number_format = number_format

            current_row += 1

        total_row = current_row
        sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=8)
        total_title = sheet.cell(row=total_row, column=1, value='Tong cong')
        total_title.font = Font(bold=True)
        total_title.alignment = Alignment(horizontal='right')

        total_quantity_cell = sheet.cell(row=total_row, column=9, value=float(total_quantity))
        total_amount_cell = sheet.cell(row=total_row, column=10, value=float(total_amount))
        total_quantity_cell.number_format = number_format
        total_amount_cell.number_format = number_format
        total_quantity_cell.font = Font(bold=True)
        total_amount_cell.font = Font(bold=True)

        column_widths = {
            'A': 7, 'B': 18, 'C': 28, 'D': 18, 'E': 20, 'F': 16, 'G': 20, 'H': 12, 'I': 16, 'J': 18,
        }
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width

        filename = f'bao_cao_don_hang_{from_date.isoformat()}_{to_date.isoformat()}.xlsx'
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        workbook.save(response)
        return response

    @staticmethod
    def export_pdf(orders, status_filter, search_query, from_date, to_date, generated_by):
        status_labels = dict(SalesOrder.STATUS_CHOICES)
        status_label = status_labels.get(status_filter, status_filter) if status_filter else 'Tat ca trang thai'
        search_label = search_query if search_query else 'Khong co'

        filename = f'bao_cao_don_hang_{from_date.isoformat()}_{to_date.isoformat()}.pdf'
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(A4),
            leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20,
        )

        styles = getSampleStyleSheet()
        regular_font = 'Helvetica'
        bold_font = 'Helvetica-Bold'

        regular_font_path = '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'
        bold_font_path = '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'
        if os.path.exists(regular_font_path) and os.path.exists(bold_font_path):
            pdfmetrics.registerFont(TTFont('DejaVuSans', regular_font_path))
            pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', bold_font_path))
            regular_font = 'DejaVuSans'
            bold_font = 'DejaVuSans-Bold'

        title_style = ParagraphStyle(
            'OrderReportTitle', parent=styles['Heading1'],
            alignment=TA_CENTER, fontName=bold_font, fontSize=16, spaceAfter=10,
        )
        body_style = ParagraphStyle(
            'OrderReportBody', parent=styles['Normal'],
            fontName=regular_font, fontSize=10,
        )

        story = [
            Paragraph('BAO CAO DON HANG', title_style),
            Paragraph(f'Tu ngay: {from_date.strftime("%d/%m/%Y")}', body_style),
            Paragraph(f'Den ngay: {to_date.strftime("%d/%m/%Y")}', body_style),
            Paragraph(f'Trang thai: {status_label}', body_style),
            Paragraph(f'Tu khoa tim kiem: {search_label}', body_style),
            Paragraph(f'Xuat luc: {timezone.localtime().strftime("%d/%m/%Y %H:%M:%S")}', body_style),
            Paragraph(f'Nguoi xuat: {generated_by}', body_style),
            Spacer(1, 10),
        ]

        table_data = [[
            'STT', 'Ma don', 'Khach hang', 'Nhan vien tao', 'Trang thai',
            'Ngay tao', 'So dong SP', 'Tong so luong', 'Tong tien',
        ]]

        total_quantity = Decimal('0')
        total_amount = Decimal('0')

        for index, order in enumerate(orders, start=1):
            order_items = list(order.items.all())
            item_count = len(order_items)
            quantity_sum = sum((item.quantity for item in order_items), Decimal('0'))
            amount_sum = sum((item.subtotal for item in order_items), Decimal('0'))

            total_quantity += quantity_sum
            total_amount += amount_sum

            table_data.append([
                str(index),
                str(order.order_code),
                str(order.customer_name),
                str(get_user_display_name(order.created_by)),
                str(order.get_status_display()),
                str(timezone.localtime(order.created_at).strftime('%d/%m/%Y %H:%M')),
                str(item_count),
                format_report_number(quantity_sum),
                format_report_number(amount_sum),
            ])

        table_data.append([
            '', 'TONG CONG', '', '', '', '', '',
            format_report_number(total_quantity), format_report_number(total_amount),
        ])

        table = Table(
            table_data,
            repeatRows=1,
            colWidths=[30, 90, 130, 115, 90, 95, 70, 90, 90],
        )
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), bold_font),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
            ('FONTNAME', (0, 1), (-1, -2), regular_font),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (6, 1), (-1, -1), 'RIGHT'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F3F6FA')),
            ('FONTNAME', (0, -1), (-1, -1), bold_font),
        ])
        table.setStyle(table_style)

        story.append(table)
        doc.build(story)
        return response
