from collections import OrderedDict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db.models import QuerySet
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from apps.order.models import SalesOrder
from apps.product.models import Product

RECEIPT_HEADERS = [
    'receipt_code',
    'product_id',
    'product_name',
    'quantity',
    'unit_price',
    'item_note',
    'receipt_note',
    'sales_order_code',
]


@dataclass
class ParsedReceiptGroup:
    receipt_code: str
    receipt_note: str
    sales_order: SalesOrder | None
    items: list


def _set_headers(worksheet, headers):
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)


def build_template_workbook(kind):
    workbook = Workbook()
    template = workbook.active
    template.title = 'Template'
    _set_headers(template, RECEIPT_HEADERS)
    example_code = 'PN-IMPORT-001' if kind == 'import' else 'EX-EXPORT-001'
    template.append(
        [
            example_code,
            '',
            'Xi măng Portland',
            10,
            50000,
            'Ghi chú dòng',
            'Ghi chú phiếu',
            '',
        ]
    )

    guide = workbook.create_sheet('HuongDan')
    guide.append(['Cột', 'Mô tả'])
    guide['A1'].font = guide['B1'].font = Font(bold=True)
    for row in [
        ('receipt_code', 'Tùy chọn. Cùng mã sẽ được gom thành một phiếu.'),
        ('product_id', 'Ưu tiên nếu có. Có thể bỏ trống nếu dùng product_name.'),
        ('product_name', 'Tên sản phẩm duy nhất trong hệ thống.'),
        ('quantity', 'Bắt buộc > 0.'),
        ('unit_price', 'Bắt buộc >= 0.'),
        ('item_note', 'Ghi chú cho từng dòng.'),
        ('receipt_note', 'Ghi chú chung cho phiếu.'),
        ('sales_order_code', 'Chỉ dùng cho phiếu xuất nếu muốn liên kết đơn hàng.'),
    ]:
        guide.append(row)
    return workbook


def _to_decimal(value, field_name, row_number, allow_zero=False):
    try:
        number = Decimal(str(value).strip())
    except (InvalidOperation, AttributeError):
        raise ValueError(f'Dòng {row_number}: {field_name} không hợp lệ.')
    if allow_zero:
        if number < 0:
            raise ValueError(f'Dòng {row_number}: {field_name} phải >= 0.')
    elif number <= 0:
        raise ValueError(f'Dòng {row_number}: {field_name} phải > 0.')
    return number


def _get_text(value):
    return str(value).strip() if value is not None else ''


def _resolve_product(product_id, product_name, row_number):
    if product_id:
        product = Product.objects.filter(pk=product_id).first()
        if product:
            return product
    if product_name:
        product = Product.objects.filter(name__iexact=product_name).first()
        if product:
            return product
    raise ValueError(f'Dòng {row_number}: không tìm thấy sản phẩm.')


def _resolve_sales_order(order_code, row_number):
    if not order_code:
        return None
    order = SalesOrder.objects.filter(order_code=order_code).first()
    if not order:
        raise ValueError(f'Dòng {row_number}: không tìm thấy đơn hàng {order_code}.')
    return order


def parse_receipt_excel(file_obj):
    try:
        workbook = load_workbook(filename=file_obj, data_only=True)
    except Exception as exc:
        raise ValueError(f'Không đọc được file Excel: {exc}') from exc

    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError('File Excel không có dữ liệu.')

    headers = [_get_text(value) for value in header_row]
    missing = [header for header in RECEIPT_HEADERS if header not in headers]
    if missing:
        raise ValueError(f'Thiếu cột bắt buộc: {", ".join(missing)}')

    index_map = {header: headers.index(header) for header in RECEIPT_HEADERS}
    grouped = OrderedDict()
    auto_group_key = '__AUTO__'

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, '') for value in row):
            continue

        receipt_code = _get_text(row[index_map['receipt_code']])
        product_id = _get_text(row[index_map['product_id']])
        product_name = _get_text(row[index_map['product_name']])
        quantity = _to_decimal(row[index_map['quantity']], 'quantity', row_number)
        unit_price = _to_decimal(row[index_map['unit_price']], 'unit_price', row_number, allow_zero=True)
        item_note = _get_text(row[index_map['item_note']])
        receipt_note = _get_text(row[index_map['receipt_note']])
        sales_order_code = _get_text(row[index_map['sales_order_code']])

        product = _resolve_product(product_id, product_name, row_number)
        sales_order = _resolve_sales_order(sales_order_code, row_number)
        group_key = receipt_code or auto_group_key

        if group_key not in grouped:
            grouped[group_key] = ParsedReceiptGroup(
                receipt_code=receipt_code,
                receipt_note=receipt_note,
                sales_order=sales_order,
                items=[],
            )
        group = grouped[group_key]
        if receipt_note and not group.receipt_note:
            group.receipt_note = receipt_note
        if sales_order and group.sales_order and group.sales_order.id != sales_order.id:
            raise ValueError(f'Dòng {row_number}: cùng receipt_code nhưng khác đơn hàng liên kết.')
        if sales_order:
            group.sales_order = sales_order

        group.items.append(
            {
                'product_id': str(product.id),
                'quantity': quantity,
                'unit_price': unit_price,
                'note': item_note,
            }
        )

    if not grouped:
        raise ValueError('File Excel không có dòng dữ liệu hợp lệ.')
    return list(grouped.values())


def export_receipts_workbook(receipts: QuerySet, kind: str):
    workbook = Workbook()
    summary = workbook.active
    summary.title = 'Phieu'
    summary_headers = [
        'receipt_code',
        'status',
        'created_by',
        'reviewed_by',
        'created_at',
        'reviewed_at',
        'receipt_note',
        'sales_order_code',
        'total_items',
    ]
    _set_headers(summary, summary_headers)

    details = workbook.create_sheet('ChiTiet')
    detail_headers = [
        'receipt_code',
        'product_id',
        'product_name',
        'quantity',
        'unit_price',
        'item_note',
        'status',
        'created_by',
        'reviewed_by',
        'created_at',
        'reviewed_at',
        'sales_order_code',
    ]
    _set_headers(details, detail_headers)

    queryset = receipts.prefetch_related('items__product').select_related('created_by', 'reviewed_by')
    if kind == 'export':
        queryset = queryset.select_related('sales_order')

    for receipt in queryset:
        sales_order_code = getattr(getattr(receipt, 'sales_order', None), 'order_code', '')
        summary.append(
            [
                receipt.receipt_code,
                receipt.get_status_display(),
                getattr(receipt.created_by, 'username', ''),
                getattr(receipt.reviewed_by, 'username', ''),
                receipt.created_at.strftime('%Y-%m-%d %H:%M:%S') if receipt.created_at else '',
                receipt.reviewed_at.strftime('%Y-%m-%d %H:%M:%S') if receipt.reviewed_at else '',
                receipt.note or '',
                sales_order_code,
                receipt.items.count(),
            ]
        )
        for item in receipt.items.all():
            details.append(
                [
                    receipt.receipt_code,
                    str(item.product_id),
                    item.product.name,
                    float(item.quantity),
                    float(item.unit_price),
                    item.note or '',
                    receipt.get_status_display(),
                    getattr(receipt.created_by, 'username', ''),
                    getattr(receipt.reviewed_by, 'username', ''),
                    receipt.created_at.strftime('%Y-%m-%d %H:%M:%S') if receipt.created_at else '',
                    receipt.reviewed_at.strftime('%Y-%m-%d %H:%M:%S') if receipt.reviewed_at else '',
                    sales_order_code,
                ]
            )
    return workbook


def workbook_to_response_bytes(workbook):
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()
