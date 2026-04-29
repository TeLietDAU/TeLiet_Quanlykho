from django.core.management import call_command
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook

from django.test import TestCase
from django.utils import timezone
from django.urls import reverse
import uuid
from datetime import timedelta
from decimal import Decimal
from io import BytesIO
from openpyxl import load_workbook
from apps.authentication.models import User
from apps.order.models import SalesOrder
from apps.product.models import Category, Product
from apps.product.serializers import ProductSerializer
from apps.warehouse.models import ImportReceipt, ImportReceiptItem, ProductStock, ExportReceipt, ExportReceiptItem
from apps.warehouse.services import ImportReceiptService, StockService, StockReportService, ExportReceiptService
from apps.warehouse.repositories import ImportReceiptRepository, ProductStockRepository, ExportReceiptRepository


class WarehouseExcelWorkflowTestCase(TestCase):
    def setUp(self):
        self.kho_user = User.objects.create_user(username='kho01', password='Kho@123', role='KHO', full_name='Kho User')
        self.ketoan_user = User.objects.create_user(username='ketoan01', password='KeToan@123', role='KE_TOAN', full_name='Ke Toan User')
        self.sale_user = User.objects.create_user(username='sale01', password='Sale@123', role='SALE', full_name='Sale User')
        self.category = Category.objects.create(name='Vat lieu')
        self.product = Product.objects.create(name='Xi mang Portland', base_price=Decimal('50000'), base_unit='Bao', category=self.category)
        self.product2 = Product.objects.create(name='Gach nung', base_price=Decimal('3000'), base_unit='Cuc', category=self.category)

    def _build_excel_file(self, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['receipt_code', 'product_id', 'product_name', 'quantity', 'unit_price', 'item_note', 'receipt_note', 'sales_order_code'])
        for row in rows:
            sheet.append(row)
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return SimpleUploadedFile('receipts.xlsx', output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_import_excel_creates_pending_import_receipt(self):
        file_obj = self._build_excel_file([
            ['', '', self.product.name, 12, 50000, 'dong 1', 'nhap lo A', ''],
            ['', '', self.product2.name, 20, 3000, 'dong 2', 'nhap lo A', ''],
        ])

        receipts = ImportReceiptService().import_receipts_from_excel(file_obj, self.kho_user)

        self.assertEqual(len(receipts), 1)
        receipt = receipts[0]
        self.assertEqual(receipt.status, 'PENDING')
        self.assertEqual(receipt.items.count(), 2)
        self.assertIsNone(ProductStock.objects.filter(product=self.product).first())

    def test_approve_imported_import_receipt_updates_stock(self):
        file_obj = self._build_excel_file([
            ['', '', self.product.name, 12, 50000, '', 'nhap lo A', ''],
        ])
        receipt = ImportReceiptService().import_receipts_from_excel(file_obj, self.kho_user)[0]

        success, _ = ImportReceiptService().approve_receipt(receipt.id, self.ketoan_user)

        self.assertTrue(success)
        stock = ProductStock.objects.get(product=self.product)
        self.assertEqual(stock.quantity, Decimal('12'))

    def test_import_excel_creates_pending_export_receipt_without_stock_change(self):
        ProductStock.objects.create(product=self.product, quantity=Decimal('40'))
        file_obj = self._build_excel_file([
            ['', '', self.product.name, 10, 50000, 'dong xuat', 'xuat lo A', ''],
        ])

        receipts = ExportReceiptService().import_receipts_from_excel(file_obj, self.kho_user)

        self.assertEqual(len(receipts), 1)
        receipt = receipts[0]
        self.assertEqual(receipt.status, 'PENDING')
        self.product.stock.refresh_from_db()
        self.assertEqual(self.product.stock.quantity, Decimal('40'))

    def test_mark_picked_attaches_photo_and_moves_linked_order_to_picked(self):
        ProductStock.objects.create(product=self.product, quantity=Decimal('40'))
        order = SalesOrder.objects.create(
            order_code='DH-20260414-001',
            customer_name='Khach A',
            customer_phone='0901234567',
            created_by=self.sale_user,
            status='WAITING',
        )
        order.items.create(product=self.product, quantity=Decimal('10'), unit_price=Decimal('50000'))
        receipt = ExportReceipt.objects.create(
            receipt_code='EX-20260420-001',
            created_by=self.kho_user,
            sales_order=order,
            status='PREPARING',
            note='xuat theo don',
        )
        receipt.items.create(product=self.product, quantity=Decimal('10'), unit_price=Decimal('50000'))
        photo = SimpleUploadedFile('picked.jpg', b'fake-image-bytes', content_type='image/jpeg')

        success, _ = ExportReceiptService().mark_as_picked(receipt.id, self.kho_user, pickup_photo=photo)

        self.assertTrue(success)
        receipt.refresh_from_db()
        order.refresh_from_db()
        self.product.stock.refresh_from_db()
        self.assertEqual(receipt.status, 'PENDING')
        self.assertTrue(receipt.stock_deducted)
        self.assertEqual(order.status, 'PICKED')
        self.assertEqual(self.product.stock.quantity, Decimal('30'))

    def test_approve_imported_export_receipt_updates_order_to_done(self):
        ProductStock.objects.create(product=self.product, quantity=Decimal('40'))
        order = SalesOrder.objects.create(
            order_code='DH-20260414-002',
            customer_name='Khach A',
            customer_phone='0901234567',
            created_by=self.sale_user,
            status='PICKED',
        )
        order.items.create(product=self.product, quantity=Decimal('10'), unit_price=Decimal('50000'))
        file_obj = self._build_excel_file([
            ['', '', self.product.name, 10, 50000, 'xuat', 'xuat theo don', order.order_code],
        ])
        receipt = ExportReceiptService().import_receipts_from_excel(file_obj, self.kho_user)[0]
        ExportReceipt.objects.filter(id=receipt.id).update(stock_deducted=True)
        self.product.stock.quantity = Decimal('30')
        self.product.stock.save(update_fields=['quantity'])

        success, _ = ExportReceiptService().approve_receipt(receipt.id, self.ketoan_user)

        self.assertTrue(success)
        order.refresh_from_db()
        self.product.stock.refresh_from_db()
        self.assertEqual(order.status, 'DONE')
        self.assertEqual(self.product.stock.quantity, Decimal('30'))

    def test_import_excel_export_receipt_moves_linked_order_to_waiting(self):
        ProductStock.objects.create(product=self.product, quantity=Decimal('40'))
        order = SalesOrder.objects.create(
            order_code='DH-20260414-003',
            customer_name='Khach B',
            customer_phone='0901234568',
            created_by=self.sale_user,
            status='CONFIRMED',
        )
        order.items.create(product=self.product, quantity=Decimal('8'), unit_price=Decimal('50000'))
        file_obj = self._build_excel_file([
            ['', '', self.product.name, 8, 50000, 'xuat', 'xuat theo don', order.order_code],
        ])

        receipt = ExportReceiptService().import_receipts_from_excel(file_obj, self.kho_user)[0]

        order.refresh_from_db()
        self.assertEqual(receipt.sales_order_id, order.id)
        self.assertEqual(order.status, 'WAITING')

    def test_product_serializer_returns_stock_fields(self):
        ProductStock.objects.create(product=self.product, quantity=Decimal('9'))

        payload = ProductSerializer(self.product).data

        self.assertEqual(payload['stock_status'], 'LOW')
        self.assertEqual(payload['stock_status_label'], 'Sắp hết')
        self.assertEqual(str(payload['stock_quantity']), '9.00')

    def test_stock_service_includes_products_without_stock_record(self):
        rows = StockService().get_all_stocks()
        row = next(item for item in rows if item['product'].id == self.product.id)
        self.assertEqual(row['quantity'], Decimal('0'))
        self.assertEqual(row['stock_status_label'], 'Ht hàng')

    def test_seed_command_creates_balanced_demo_data(self):
        call_command('seed_inventory_demo')

        self.assertEqual(SalesOrder.objects.count(), 20)
        self.assertEqual(ImportReceipt.objects.count(), 15)
        self.assertEqual(ExportReceipt.objects.count(), 15)
        
        # Nhập kho cho cả 2
        items1 = [{'product_id': str(self.product.id), 'quantity': Decimal('100'), 'unit_price': Decimal('50000')}]
        receipt1, _ = import_service.create_receipt('Test', items1, self.kho_user)
        import_service.approve_receipt(receipt1.id, self.ketoan_user)
        
        items2 = [{'product_id': str(product2.id), 'quantity': Decimal('5000'), 'unit_price': Decimal('3000')}]
        receipt2, _ = import_service.create_receipt('Test', items2, self.kho_user)
        import_service.approve_receipt(receipt2.id, self.ketoan_user)
        
        # Lấy tất cả tồn kho
        all_stocks = self.service.get_all_stocks()
        self.assertEqual(all_stocks.count(), 2)


class StockReportServiceTestCase(TestCase):
    """Test báo cáo tồn kho theo thời gian (US-23)."""

    def setUp(self):
        self.user = User.objects.create_user(username='admin01', password='Admin@123', role='ADMIN')

        self.cat_a = Category.objects.create(name='Nhóm A')
        self.cat_b = Category.objects.create(name='Nhóm B')

        self.product_a = Product.objects.create(
            name='Sản phẩm A',
            base_price=Decimal('10000'),
            base_unit='Bao',
            category=self.cat_a,
        )
        self.product_b = Product.objects.create(
            name='Sản phẩm B',
            base_price=Decimal('20000'),
            base_unit='Bao',
            category=self.cat_b,
        )

        self.service = StockReportService()

    def _create_import(self, product, quantity, reviewed_at):
        receipt = ImportReceipt.objects.create(
            receipt_code=f'PN-TEST-{uuid.uuid4().hex[:10]}',
            created_by=self.user,
            reviewed_by=self.user,
            status='APPROVED',
            reviewed_at=reviewed_at,
        )
        ImportReceiptItem.objects.create(
            receipt=receipt,
            product=product,
            quantity=Decimal(str(quantity)),
            unit_price=Decimal('0'),
        )

    def _create_export(self, product, quantity, reviewed_at):
        receipt = ExportReceipt.objects.create(
            receipt_code=f'EX-TEST-{uuid.uuid4().hex[:10]}',
            created_by=self.user,
            reviewed_by=self.user,
            status='APPROVED',
            reviewed_at=reviewed_at,
        )
        ExportReceiptItem.objects.create(
            receipt=receipt,
            product=product,
            quantity=Decimal(str(quantity)),
            unit_price=Decimal('0'),
        )

    def test_build_report_by_time_range(self):
        today = timezone.localdate()
        from_date = today - timedelta(days=3)
        to_date = today

        before_period = timezone.now() - timedelta(days=4)
        in_period = timezone.now() - timedelta(days=1)

        # Lịch sử sản phẩm A:
        # Trước kỳ: +100, -30 => tồn đầu = 70
        # Trong kỳ: +20, -10 => tồn cuối = 80
        self._create_import(self.product_a, 100, before_period)
        self._create_export(self.product_a, 30, before_period)
        self._create_import(self.product_a, 20, in_period)
        self._create_export(self.product_a, 10, in_period)

        rows, totals = self.service.build_report(from_date, to_date)

        row_a = next(row for row in rows if row['product'].id == self.product_a.id)
        self.assertEqual(row_a['opening'], Decimal('70'))
        self.assertEqual(row_a['import_qty'], Decimal('20'))
        self.assertEqual(row_a['export_qty'], Decimal('10'))
        self.assertEqual(row_a['closing'], Decimal('80'))

        self.assertEqual(totals['opening'], sum((r['opening'] for r in rows), Decimal('0')))
        self.assertEqual(totals['closing'], sum((r['closing'] for r in rows), Decimal('0')))

    def test_build_report_with_category_filter(self):
        today = timezone.localdate()
        from_date = today - timedelta(days=7)
        to_date = today
        in_period = timezone.now() - timedelta(days=1)

        self._create_import(self.product_a, 50, in_period)
        self._create_import(self.product_b, 70, in_period)

        rows, _ = self.service.build_report(from_date, to_date, category_id=str(self.cat_a.id))

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['product'].id, self.product_a.id)


class StockReportExportExcelViewTestCase(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='admin_export', password='Admin@123', role='ADMIN')
        self.cat_a = Category.objects.create(name='Cat A')
        self.cat_b = Category.objects.create(name='Cat B')
        self.product_a = Product.objects.create(
            name='Vat tu A',
            base_price=Decimal('10000'),
            base_unit='Bao',
            category=self.cat_a,
        )
        self.product_b = Product.objects.create(
            name='Vat tu B',
            base_price=Decimal('20000'),
            base_unit='Bao',
            category=self.cat_b,
        )

    def _create_import(self, product, quantity, reviewed_at):
        receipt = ImportReceipt.objects.create(
            receipt_code=f'PN-EXCEL-{uuid.uuid4().hex[:8]}',
            created_by=self.user,
            reviewed_by=self.user,
            status='APPROVED',
            reviewed_at=reviewed_at,
        )
        ImportReceiptItem.objects.create(
            receipt=receipt,
            product=product,
            quantity=Decimal(str(quantity)),
            unit_price=Decimal('0'),
        )

    def test_export_excel_returns_xlsx_file(self):
        self.client.login(username='admin_export', password='Admin@123')
        reviewed_at = timezone.now() - timedelta(days=1)
        self._create_import(self.product_a, 40, reviewed_at)

        today = timezone.localdate()
        from_date = (today - timedelta(days=7)).isoformat()
        to_date = today.isoformat()

        response = self.client.get(
            reverse('reports:stock_report_export_excel'),
            {'from_date': from_date, 'to_date': to_date}
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', response['Content-Type'])
        self.assertIn('.xlsx', response['Content-Disposition'])

        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active

        self.assertEqual(sheet['A1'].value, 'BAO CAO TON KHO THEO THOI GIAN')
        self.assertEqual(sheet['A6'].value, 'Nguoi xuat: admin_export')
        self.assertEqual(sheet['B8'].value, 'Vat tu A')
        self.assertEqual(sheet['E8'].value, 0)
        self.assertEqual(sheet['F8'].value, 40)
        self.assertEqual(sheet['H8'].value, 40)
        self.assertEqual(sheet['F8'].number_format, '#,##0.##')

    def test_export_excel_applies_category_filter(self):
        self.client.login(username='admin_export', password='Admin@123')
        reviewed_at = timezone.now() - timedelta(days=1)
        self._create_import(self.product_a, 10, reviewed_at)
        self._create_import(self.product_b, 20, reviewed_at)

        today = timezone.localdate()
        from_date = (today - timedelta(days=7)).isoformat()
        to_date = today.isoformat()

        response = self.client.get(
            reverse('reports:stock_report_export_excel'),
            {
                'from_date': from_date,
                'to_date': to_date,
                'category': str(self.cat_a.id),
            }
        )

        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active

        # Chỉ còn 1 dòng dữ liệu của danh mục Cat A.
        self.assertEqual(sheet['B8'].value, 'Vat tu A')
        self.assertIsNone(sheet['B9'].value)

    def test_stock_report_view_blocks_invalid_date_range(self):
        self.client.login(username='admin_export', password='Admin@123')
        reviewed_at = timezone.now() - timedelta(days=1)
        self._create_import(self.product_a, 15, reviewed_at)

        today = timezone.localdate()
        response = self.client.get(
            reverse('reports:stock_report'),
            {
                'from_date': (today + timedelta(days=1)).isoformat(),
                'to_date': today.isoformat(),
            }
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['rows'], [])
        self.assertEqual(response.context['totals']['opening'], 0)
        self.assertEqual(response.context['totals']['import_qty'], 0)
        self.assertEqual(response.context['totals']['export_qty'], 0)
        self.assertEqual(response.context['totals']['closing'], 0)

    def test_export_excel_blocks_invalid_date_range(self):
        self.client.login(username='admin_export', password='Admin@123')
        today = timezone.localdate()

        response = self.client.get(
            reverse('reports:stock_report_export_excel'),
            {
                'from_date': (today + timedelta(days=1)).isoformat(),
                'to_date': today.isoformat(),
            }
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('reports:stock_report'), response['Location'])


class StockReportExportPdfViewTestCase(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='admin_pdf', password='Admin@123', role='ADMIN')
        self.category = Category.objects.create(name='Cat PDF')
        self.product = Product.objects.create(
            name='Vat tu PDF',
            base_price=Decimal('15000'),
            base_unit='Bao',
            category=self.category,
        )

    def _create_import(self, quantity, reviewed_at):
        receipt = ImportReceipt.objects.create(
            receipt_code=f'PN-PDF-{uuid.uuid4().hex[:8]}',
            created_by=self.user,
            reviewed_by=self.user,
            status='APPROVED',
            reviewed_at=reviewed_at,
        )
        ImportReceiptItem.objects.create(
            receipt=receipt,
            product=self.product,
            quantity=Decimal(str(quantity)),
            unit_price=Decimal('0'),
        )

    def test_export_pdf_returns_pdf_file(self):
        self.client.login(username='admin_pdf', password='Admin@123')
        self._create_import(25, timezone.now() - timedelta(days=1))

        today = timezone.localdate()
        response = self.client.get(
            reverse('reports:stock_report_export_pdf'),
            {
                'from_date': (today - timedelta(days=7)).isoformat(),
                'to_date': today.isoformat(),
            }
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('application/pdf', response['Content-Type'])
        self.assertIn('.pdf', response['Content-Disposition'])
        self.assertTrue(response.content.startswith(b'%PDF'))

    def test_export_pdf_blocks_invalid_date_range(self):
        self.client.login(username='admin_pdf', password='Admin@123')
        today = timezone.localdate()

        response = self.client.get(
            reverse('reports:stock_report_export_pdf'),
            {
                'from_date': (today + timedelta(days=1)).isoformat(),
                'to_date': today.isoformat(),
            }
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('reports:stock_report'), response['Location'])
