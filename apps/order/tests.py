from decimal import Decimal

from django.core.files.uploadedfile import SimpleUploadedFile

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from datetime import timedelta
from io import BytesIO
from decimal import Decimal
from openpyxl import load_workbook
from apps.authentication.models import User
from apps.product.models import Category, Product
from apps.warehouse.models import ProductStock, ExportReceipt, ExportReceiptItem
from apps.warehouse.repositories import ProductStockRepository, ExportReceiptRepository
from apps.warehouse.services import ImportReceiptService, ExportReceiptService
from apps.order.models import SalesOrder, SalesOrderItem
from apps.order.services import SalesOrderService


class SalesOrderWorkflowTestCase(TestCase):
    def setUp(self):
        self.sale_user = User.objects.create_user(username='sale01', password='Sale@123', role='SALE', full_name='Sale User')
        self.admin_user = User.objects.create_user(username='admin01', password='Admin@123', role='ADMIN', is_superuser=True, is_staff=True, full_name='Admin User')
        self.kho_user = User.objects.create_user(username='kho01', password='Kho@123', role='KHO', full_name='Kho User')
        self.ketoan_user = User.objects.create_user(username='ketoan01', password='KeToan@123', role='KE_TOAN', full_name='Ke Toan User')
        self.category = Category.objects.create(name='Vat lieu xay dung')
        self.product = Product.objects.create(name='Thep cay D16', base_price=Decimal('185000'), base_unit='Cay', category=self.category)
        ProductStock.objects.create(product=self.product, quantity=Decimal('100'))
        self.sales_service = SalesOrderService()
        self.export_service = ExportReceiptService()

    def _create_order(self):
        order, errors = self.sales_service.create_order(
            customer_name='Cong ty ABC',
            customer_phone='0901234567',
            note='don test',
            items_data=[{'product_id': str(self.product.id), 'quantity': Decimal('10'), 'unit_price': Decimal('185000')}],
            user=self.sale_user,
        )
        self.assertIsNone(errors)
        return order

    def test_update_to_waiting_creates_linked_export_receipt_in_preparing(self):
        order = self._create_order()

        success, _ = self.sales_service.update_status(order.id, 'WAITING', updated_by=self.admin_user)

        self.assertTrue(success)
        receipt = ExportReceipt.objects.get(sales_order=order)
        self.assertEqual(receipt.status, 'PREPARING')
        self.assertEqual(receipt.items.count(), 1)

    def test_mark_picked_moves_order_to_picked_and_deducts_stock(self):
        order = self._create_order()
        self.sales_service.update_status(order.id, 'WAITING', updated_by=self.admin_user)
        receipt = ExportReceipt.objects.get(sales_order=order)
        photo = SimpleUploadedFile('picked.jpg', b'fake-image-bytes', content_type='image/jpeg')

        success, _ = self.export_service.mark_as_picked(receipt.id, self.kho_user, pickup_photo=photo)

        self.assertTrue(success)
        receipt.refresh_from_db()
        order.refresh_from_db()
        self.product.stock.refresh_from_db()
        self.assertEqual(receipt.status, 'PENDING')
        self.assertTrue(receipt.stock_deducted)
        self.assertTrue(bool(receipt.pickup_photo))
        self.assertEqual(order.status, 'PICKED')
        self.assertEqual(self.product.stock.quantity, Decimal('90'))

    def test_approve_after_pick_moves_order_to_done_without_second_stock_deduction(self):
        order = self._create_order()
        self.sales_service.update_status(order.id, 'WAITING', updated_by=self.admin_user)
        receipt = ExportReceipt.objects.get(sales_order=order)
        self.export_service.mark_as_picked(receipt.id, self.kho_user)

        success, _ = self.export_service.approve_receipt(receipt.id, self.ketoan_user)

        self.assertTrue(success)
        order.refresh_from_db()
        self.product.stock.refresh_from_db()
        self.assertEqual(order.status, 'DONE')
        self.assertEqual(self.product.stock.quantity, Decimal('90'))

    def test_reject_after_pick_restores_stock_and_returns_order_to_waiting(self):
        order = self._create_order()
        self.sales_service.update_status(order.id, 'WAITING', updated_by=self.admin_user)
        receipt = ExportReceipt.objects.get(sales_order=order)
        self.export_service.mark_as_picked(receipt.id, self.kho_user)

        success, _ = self.export_service.reject_receipt(receipt.id, self.ketoan_user, 'thieu anh')

        self.assertTrue(success)
        order.refresh_from_db()
        self.product.stock.refresh_from_db()
        receipt.refresh_from_db()
        self.assertEqual(order.status, 'WAITING')
        self.assertEqual(receipt.status, 'REJECTED')
        self.assertEqual(self.product.stock.quantity, Decimal('100'))

    def test_cancel_order_returns_stock_when_goods_already_picked(self):
        order = self._create_order()
        self.sales_service.update_status(order.id, 'WAITING', updated_by=self.admin_user)
        receipt = ExportReceipt.objects.get(sales_order=order)
        self.export_service.mark_as_picked(receipt.id, self.kho_user)

        success, _ = self.sales_service.update_status(order.id, 'CANCELLED')

        self.assertTrue(success)
        self.product.stock.refresh_from_db()
        receipt.refresh_from_db()
        self.assertEqual(self.product.stock.quantity, Decimal('100'))
        self.assertEqual(receipt.status, 'REJECTED')

    def test_invalid_transition_is_rejected(self):
        order = self._create_order()

        success, message = self.sales_service.update_status(order.id, 'DONE')

        self.assertFalse(success)
        self.assertIn('Khong the chuyen', message)

    def test_mark_picked_is_blocked_when_stock_is_insufficient(self):
        order = self._create_order()
        self.sales_service.update_status(order.id, 'WAITING', updated_by=self.admin_user)
        receipt = ExportReceipt.objects.get(sales_order=order)
        self.product.stock.quantity = Decimal('5')
        self.product.stock.save(update_fields=['quantity'])

        success, message = self.export_service.mark_as_picked(receipt.id, self.kho_user)

        self.assertFalse(success)
        self.assertIn('ton kho khong du', message)
        receipt.refresh_from_db()
        order.refresh_from_db()
        self.assertEqual(receipt.status, 'PREPARING')
        self.assertEqual(order.status, 'WAITING')
        
        # Phiếu xuất được tạo tự động
        export = ExportReceipt.objects.get(note__icontains=order.order_code)
        self.assertEqual(export.status, 'PENDING')
        
        # ===== STEP 4: DUYỆT PHIẾU XUẤT =====
        success, msg = export_service.approve_receipt(export.id, self.ketoan_user)
        self.assertTrue(success)
        
        # Tồn kho = 150 (200 - 50)
        stock = ProductStockRepository.get_stock(self.product.id)
        self.assertEqual(stock.quantity, Decimal('150'))
        
        # Đơn hàng → DONE
        order.refresh_from_db()
        self.assertEqual(order.status, 'DONE')
        
        # Phiếu xuất → APPROVED
        export.refresh_from_db()
        self.assertEqual(export.status, 'APPROVED')


class SalesOrderExportExcelViewTestCase(TestCase):
    def setUp(self):
        self.admin_user = User.objects.create_user(
            username='admin_export_order',
            password='Admin@123',
            role='ADMIN',
            is_staff=True,
            is_superuser=True,
        )
        self.sale_user = User.objects.create_user(
            username='sale_export_order',
            password='Sale@123',
            role='SALE',
        )

        self.category = Category.objects.create(name='Danh muc export')
        self.product_a = Product.objects.create(
            name='San pham A',
            base_price=Decimal('50000'),
            base_unit='Bao',
            category=self.category,
        )
        self.product_b = Product.objects.create(
            name='San pham B',
            base_price=Decimal('30000'),
            base_unit='Thung',
            category=self.category,
        )

        now = timezone.now()
        self.order_confirmed = self._create_order(
            order_code='DH-EXCEL-001',
            customer_name='Cong ty A',
            status='CONFIRMED',
            created_at=now - timedelta(days=3),
        )
        self.order_done = self._create_order(
            order_code='DH-EXCEL-002',
            customer_name='Cong ty B',
            status='DONE',
            created_at=now - timedelta(days=1),
        )

    def _create_order(self, order_code, customer_name, status, created_at):
        order = SalesOrder.objects.create(
            order_code=order_code,
            customer_name=customer_name,
            customer_phone='0900000000',
            created_by=self.sale_user,
            status=status,
        )
        SalesOrderItem.objects.create(
            order=order,
            product=self.product_a,
            quantity=Decimal('2'),
            unit_price=Decimal('50000'),
        )
        SalesOrderItem.objects.create(
            order=order,
            product=self.product_b,
            quantity=Decimal('1.5'),
            unit_price=Decimal('30000'),
        )
        SalesOrder.objects.filter(id=order.id).update(created_at=created_at)
        order.refresh_from_db()
        return order

    def test_export_excel_returns_xlsx_file(self):
        self.client.login(username='admin_export_order', password='Admin@123')
        today = timezone.localdate()
        response = self.client.get(
            reverse('reports:sales_export_excel'),
            {
                'from_date': (today - timedelta(days=7)).isoformat(),
                'to_date': today.isoformat(),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', response['Content-Type'])
        self.assertIn('.xlsx', response['Content-Disposition'])

        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active
        self.assertEqual(sheet['A1'].value, 'BAO CAO DON HANG')

        order_codes = [
            sheet.cell(row=row, column=2).value
            for row in range(10, sheet.max_row)
            if sheet.cell(row=row, column=2).value
        ]
        self.assertIn(self.order_confirmed.order_code, order_codes)
        self.assertIn(self.order_done.order_code, order_codes)
        self.assertEqual(sheet['I10'].number_format, '#,##0.##')
        self.assertEqual(sheet['J10'].number_format, '#,##0.##')

    def test_export_excel_applies_status_filter(self):
        self.client.login(username='admin_export_order', password='Admin@123')
        today = timezone.localdate()
        response = self.client.get(
            reverse('reports:sales_export_excel'),
            {
                'status': 'DONE',
                'from_date': (today - timedelta(days=7)).isoformat(),
                'to_date': today.isoformat(),
            },
        )

        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active
        order_codes = [
            sheet.cell(row=row, column=2).value
            for row in range(10, sheet.max_row)
            if sheet.cell(row=row, column=2).value
        ]

        self.assertIn(self.order_done.order_code, order_codes)
        self.assertNotIn(self.order_confirmed.order_code, order_codes)

    def test_export_excel_blocks_invalid_date_range(self):
        self.client.login(username='admin_export_order', password='Admin@123')
        today = timezone.localdate()

        response = self.client.get(
            reverse('reports:sales_export_excel'),
            {
                'from_date': (today + timedelta(days=1)).isoformat(),
                'to_date': today.isoformat(),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('order:sales_list'), response['Location'])


class SalesOrderExportPdfViewTestCase(TestCase):
    def setUp(self):
        self.admin_user = User.objects.create_user(
            username='admin_pdf_order',
            password='Admin@123',
            role='ADMIN',
            is_staff=True,
            is_superuser=True,
        )
        self.sale_user = User.objects.create_user(
            username='sale_pdf_order',
            password='Sale@123',
            role='SALE',
        )

        self.category = Category.objects.create(name='Danh muc PDF')
        self.product = Product.objects.create(
            name='San pham PDF',
            base_price=Decimal('45000'),
            base_unit='Bao',
            category=self.category,
        )

        now = timezone.now()
        self.order_done = SalesOrder.objects.create(
            order_code='DH-PDF-001',
            customer_name='Cong ty PDF',
            customer_phone='0901231234',
            created_by=self.sale_user,
            status='DONE',
        )
        SalesOrderItem.objects.create(
            order=self.order_done,
            product=self.product,
            quantity=Decimal('3'),
            unit_price=Decimal('45000'),
        )
        SalesOrder.objects.filter(id=self.order_done.id).update(created_at=now - timedelta(days=1))

    def test_export_pdf_returns_pdf_file(self):
        self.client.login(username='admin_pdf_order', password='Admin@123')
        today = timezone.localdate()

        response = self.client.get(
            reverse('reports:sales_export_pdf'),
            {
                'from_date': (today - timedelta(days=7)).isoformat(),
                'to_date': today.isoformat(),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('application/pdf', response['Content-Type'])
        self.assertIn('.pdf', response['Content-Disposition'])
        self.assertTrue(response.content.startswith(b'%PDF'))

    def test_export_pdf_blocks_invalid_date_range(self):
        self.client.login(username='admin_pdf_order', password='Admin@123')
        today = timezone.localdate()

        response = self.client.get(
            reverse('reports:sales_export_pdf'),
            {
                'from_date': (today + timedelta(days=1)).isoformat(),
                'to_date': today.isoformat(),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('order:sales_list'), response['Location'])

